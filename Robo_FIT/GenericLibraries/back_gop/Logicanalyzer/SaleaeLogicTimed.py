from saleae import automation
#from saleae import Saleae
import os.path
from datetime import datetime
import openpyxl
import csv
from Robo_FIT.GenericLibraries.GenericOpLibs.Logicanalyzer.ConfigurationManager import ConfigurationManager
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass


class SaleaeLogicTimed:
    def __init__(self):
        self.config_manager = ConfigurationManager()
        self.common_keyword = CommonKeywordsClass()

    def logic_timed(self, duration):
        """
        This method is used to run TimedCaptureMode in saleae . This record waveform , Calculate Duty cycle & frequency.
        :param : duration - Duration to stop capture in seconds
        :return:digital_path - Path of digital.csv export.
        """

        with automation.Manager.connect(port=self.config_manager.get_port()) as manager:
            """
            The settings chosen here will depend on your device's capabilities and what
            you can configure in the Logic 2 UI.
            """
            enabled_digital_channels1, digital_sample_rate1, digital_threshold_volts1 = self.config_manager.get_device_config()
            device_configuration = automation.LogicDeviceConfiguration(
                enabled_digital_channels=enabled_digital_channels1,
                digital_sample_rate=int(digital_sample_rate1),
                digital_threshold_volts=float(digital_threshold_volts1),
            )
            print("Device config =", device_configuration)

            """
                Stop recording after "duration" seconds
            """
            capture_configuration = automation.CaptureConfiguration(
                capture_mode=automation.TimedCaptureMode(duration_seconds=int(duration))
            )
            """
             Start a capture - the capture will be automatically closed when leaving the `with` block
             Note: The serial number 'F4241' is for the Logic Pro 16 demo device.
                   To use a real device, you can:
                     1. Omit the `device_id` argument. Logic 2 will choose the first real (non-simulated) device.
                     2. Use the serial number for your device. See the "Finding the Serial Number
                        of a Device" section for information on finding your device's serial number.
            """
            with manager.start_capture(
                    device_id=self.config_manager.get_device_id(),
                    device_configuration=device_configuration,
                    capture_configuration=capture_configuration) as capture:
                capture.wait()
                #output_dir = os.path.join(os.getcwd(), f'output-{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}')
                output_dir = os.path.join(self.common_keyword.get_logic_report_path(),
                                          f'output-{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}')
                os.makedirs(output_dir)
                capture.export_raw_data_csv(directory=output_dir, digital_channels=enabled_digital_channels1)
                digital_path = os.path.join(output_dir, 'digital.csv')
                capture_filepath = os.path.join(output_dir, 'example_capture.sal')
                capture.save_capture(filepath=capture_filepath)
                #self.dcycle(digital_path)
                return digital_path

    def dcycle_freq(self, El_path):

        """
            This method can be called only after getting digital path from logic() function
            :param : El_path - Path of digital.csv export
            :return:Dutycyle - Dutycyle & Freq - Frequency
        """
        time_val = []
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(El_path) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
            wb.save('output.xlsx')
        output_dir = os.path.join(os.getcwd(), 'output.xlsx')
        wb_obj = openpyxl.load_workbook(output_dir)
        sheet_obj = wb_obj.active
        for i in range(0, 3):
            cell_obj = sheet_obj.cell(row=i + 3, column=1)
            temp = cell_obj.value
            try:
                temp = float(temp)
                time_val.append(temp)
            except Exception as exp_err:
                print("Not enough inputs available in output.xlsx ", exp_err)
        try:
            print("Time val 0 = ", time_val[0], "Time val 1 = ", time_val[1], "Time val 2 = ", time_val[2])
            duration1 = time_val[1] - time_val[0]
            duration1 = float(duration1)
            duration2 = time_val[2] - time_val[1]
            duration2 = float(duration2)
            duty_cycle = duration1 / (duration1 + duration2)
            print('Duty Cycle =', duty_cycle)
            print("Duration1 = ", duration1, "Duration2 = ", duration2)
            frequency = 1 / (duration1 + duration2)
            print(frequency, ' Hz')
            return duty_cycle, frequency
        except Exception as exp_err:
            print("Not enough inputs available in output.xlsx . Exception = ", exp_err)

#if __name__ == '__main__':
#    h = SaleaeLogicTimed()
#    dpath = h.logic_timed(5)
#    print("Digital path = ", dpath)
#    dcycle, freq = h.dcycle_freq(dpath)
#    print("Dcycle & Freq = ", dcycle, freq)
