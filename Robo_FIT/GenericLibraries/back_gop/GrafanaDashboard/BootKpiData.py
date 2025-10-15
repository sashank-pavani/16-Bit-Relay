import openpyxl
import json
from Robo_FIT.GenericLibraries.GenericOpLibs.GrafanaDashboard.DataPost import DataPost
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.Constants import *
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.ProjectConfigManager import ProjectConfigManager
import os
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import robot_print_error, robot_print_info


class BootKpiData:

    def __init__(self):
        self.project_config_file = ProjectConfigManager()
        self.base_url = self.project_config_file.get_grafana_base_url()
        self.helper = CommonKeywordsClass()
        self.upload_to_grafana = DataPost()
        self.workbook = openpyxl.load_workbook(self.excel_file_path)

    @property
    def excel_file_path(self):
        """
        This function gets excel file path of BOOT KPI
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_BOOT_KPI_DIR_NAME,
            ROBO_BOOT_KPI_EXCEL_NAME
        )
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found at path: {file_path}")
        return file_path

    @property
    def upload_url(self):
        """
        This function returns the API endpoint URL for BOOT KPI
        :return: URL
        """
        return f"{self.base_url}{BOOTKPI_URL}"

    def _check_bootkpi_file(self):
        """
        This function checks whether excel file path exists, if it does, it checks that sheet is not blank
        :return: True if excel file exists and is not blank else false
        """
        path = self.excel_file_path
        sheet = self._get_sheet_name()
        if sheet is None:
            robot_print_error("The sheet is empty")
            return False
        if os.path.exists(path):
            if self._is_sheet_not_blank(sheet):
                robot_print_info(f"File generated and sheet is present for Boot KPI", print_in_report=True,
                                 underline=True)
                return True
        else:
            robot_print_info("No file generated for Boot KPI")
            return False

    def _is_sheet_not_blank(self, sheet_name):
        """
       This function checks whether the excel sheet contains data or not.
    :return: True if the sheet is not blank, False otherwise.
        """
        try:
            sheet = self.workbook[sheet_name]
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                if row_index == 1:
                    continue
                for cell in row:
                    if cell.value is not None:
                        self.workbook.close()
                        return True
            self.workbook.close()
            robot_print_info("Sheet does not contain any data.")
            return False
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")
            return False

    def _get_sheet_name(self):
        """
        This function gets name of sheet in excel file
        :return: Sheet name
        """
        try:
            excel_file_path = self.excel_file_path
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet_names = workbook.sheetnames
            workbook.close()
            for sheet_name in sheet_names:
                if "boot_kpi" in sheet_name.lower():
                    robot_print_info(f"Found 'boot_kpi' sheet: {sheet_name}")
                    return sheet_name

            robot_print_info("No 'boot_kpi' sheet found")
            return None
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")

    def upload_bootkpi_data(self):
        """
        This function prepares data and calls post function to post BOOT KPI data to grafana dashboard
        """
        try:
            if self._check_bootkpi_file():
                bootkpi_excel_file = self.excel_file_path
                bootkpi_sheet_name = self._get_sheet_name()
                data_json = {}
                data_json_list = self.upload_to_grafana.convert_excel_to_json(bootkpi_excel_file, bootkpi_sheet_name)
                if isinstance(data_json_list, list):
                    if data_json_list:
                        for item in data_json_list:
                            data_json[str(item.get('Iteration'))] = item
                        data_value = json.dumps(data_json)
                    else:
                        raise ValueError("Empty list received from convert_excel_to_json")
                elif isinstance(data_json_list, str):
                    data_value = data_json_list
                else:
                    raise ValueError("Unexpected data type for data_json_list")

                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_value
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Boot KPI data upload: {e}")
