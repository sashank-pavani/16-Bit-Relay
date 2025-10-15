import openpyxl

from Robo_FIT.GenericLibraries.GenericOpLibs.GrafanaDashboard.DataPost import DataPost
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.Constants import *
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.ProjectConfigManager import ProjectConfigManager
import os
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import robot_print_error, robot_print_info


class CpuUtilisationData:
    def __init__(self):

        self.project_config_file = ProjectConfigManager()
        self.base_url = self.project_config_file.get_grafana_base_url()
        self.helper = CommonKeywordsClass()
        self.upload_to_grafana = DataPost()
        self.base_url = self.project_config_file.get_grafana_base_url()

    @property
    def android_cpu_excel_file_path(self):
        """
        This function gets excel file path of Android CPU
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_CPU_DIR_NAME,
            ROBO_ANDROID_CPU_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_cpu_excel_file_path(self):
        """
        This function gets excel file path of QNX CPU
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_CPU_DIR_NAME,
            ROBO_QNX_CPU_EXCEL_NAME
        )
        return file_path

    @property
    def upload_url(self):
        """
        This function returns the API endpoint URL for CPU Utilisation
        :return: URL
        """
        return self.base_url + CPU_URL

    def _check_cpu_file(self, file_type: str):
        """
        This function checks whether excel file path exists, if it does, it checks that sheet is not blank
        :return: True if excel file exists and is not blank else false
        """
        if file_type == ROBO_ANDROID_CPU_FILE_NAME:
            path = self.android_cpu_excel_file_path
            sheet = self._get_sheet_name(ROBO_ANDROID_CPU_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Boot KPI", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Boot KPI")
                return False
        else:
            path = self.qnx_cpu_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_CPU_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Boot KPI", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Boot KPI")
                return False

    def _is_sheet_not_blank(self, excel_file_path, sheet_name):
        """
        This function checks that excel file is not blank
        :return:True if file is not blank else false
        """
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        workbook.close()
                        print("File is not empty")
                        return True
            workbook.close()
            robot_print_info("File is empty")
            return False
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")
            return False

    def _get_sheet_name(self, file_type: str):
        """
        This function gets name of sheet in excel file
        :return: Sheet name
        """
        try:
            if file_type == ROBO_ANDROID_CPU_FILE_NAME:
                excel_file_path = self.android_cpu_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            else:
                excel_file_path = self.qnx_cpu_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")

    def __upload_android_cpu_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_cpu_file(ROBO_ANDROID_CPU_FILE_NAME):
                android_cpu_data_excel_file = self.android_cpu_excel_file_path
                android_cpu_data_sheet_name = self._get_sheet_name(android_cpu_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(android_cpu_data_excel_file, android_cpu_data_sheet_name)

                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Android CPU data upload: {e}")

    def __upload_qnx_cpu_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_cpu_file(ROBO_QNX_CPU_FILE_NAME):
                qnx_cpu_data_excel_file = self.android_cpu_excel_file_path
                qnx_cpu_data_sheet_name = self._get_sheet_name(qnx_cpu_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_cpu_data_excel_file,
                                                                         qnx_cpu_data_sheet_name)

                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json

                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during QNX CPU data upload: {e}")