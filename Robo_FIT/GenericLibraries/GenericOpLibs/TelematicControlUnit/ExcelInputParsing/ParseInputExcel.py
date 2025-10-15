import json
import pandas as pd
from typing import Dict, List
from openpyxl import load_workbook
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import robot_print_error


class ParseInputExcel:
    # key data
    _TOPIC_NAME = "topicName"
    _INPUT_MESSAGE = "inMsg"
    _EXPECTED_OUTPUT = "exptOutput"
    _TEST_OUTPUT = "testOutput"
    _KIP_VALUE = "kpiVal"

    def __init__(self, input_file_path, sheet_name=0):
        """
        This class is parse the input excel file and provide the according to user need.
        :param input_file_path: path of the Excel file.
        """
        self.input_file_path = input_file_path
        self.input_sheet = pd.read_excel(input_file_path, sheet_name=sheet_name)

    def _get_excel_output_dict(self) -> dict:
        """
        This method read the data/records for the In[put Excel file and convert them into Dictionary
        :return: all record dictionary
        Dictionary as like:
            input_dict = {
                            "testCaseName1"{
                                "topicName" : "",
                                "inMsg" : "",
                                "exptOutput" : "",
                                "testOutput" : "",
                                "kpiVal" : ""
                            },
                            "testCaseName2"{
                                "topicName" : "",
                                "inMsg" : "",
                                "exptOutput" : "",
                                "testOutput" : "",
                                "kpiVal" : ""
                            },
                        }
        """
        # get all the data/records form the excel file
        tn = self.input_sheet.set_index("Test_Step_Description")["Topic_Name"].to_dict()
        ti = self.input_sheet.set_index("Test_Step_Description")["Input_Message"].to_dict()
        ex = self.input_sheet.set_index("Test_Step_Description")["Expected_Output"].to_dict()
        to = self.input_sheet.set_index("Test_Step_Description")["Test_Output"].to_dict()
        kip = self.input_sheet.set_index("Test_Step_Description")["KPI"].to_dict()

        # create a list of dict data
        input_list = [tn, ti, ex, to, kip]
        # initialize dict to store final output
        input_dict = {}
        # iterate all the records and create a Dictionary
        for k in tn.keys():
            input_dict[k] = tuple(d[k] for d in input_list)
            input_dict[k] = {
                ParseInputExcel._TOPIC_NAME: input_dict[k][0],
                ParseInputExcel._INPUT_MESSAGE: input_dict[k][1],
                ParseInputExcel._EXPECTED_OUTPUT: input_dict[k][2],
                ParseInputExcel._TEST_OUTPUT: input_dict[k][3],
                ParseInputExcel._KIP_VALUE: input_dict[k][4]
            }
        # return the final dict
        return input_dict

    def get_test_case_list(self) -> List:
        """
        This method provide the List of all the test case names available into excel file.
        :return: List of test case names
        """
        return list(self.input_sheet["Test_Step_Description"])

    def get_topic_name(self, test_description) -> str:
        """
        This method provide the topic name form the excel file corresponding to the give test case name
        :param test_description: test case name who's topic name user want.
        :return: Topic name corresponding to given test case name
        """
        try:
            if test_description in self.get_test_case_list():
                excel_dict = self._get_excel_output_dict()
                return excel_dict[test_description][ParseInputExcel._TOPIC_NAME]
            else:
                robot_print_error(f"The given Test case name is not match in input excel file...!!!",
                                  print_in_report=True)
        except IOError as ioerr:
            robot_print_error(f"The given Test case name is not match in InputExcel...!!!"
                              f"\nPlease check the test case name...!!!\nEXCEPTION: {ioerr}",
                              print_in_report=True)

    def get_input_message(self, test_description) -> json:
        """
        This method provide the CAN Message form the excel file corresponding to the give test case name
        :param test_description: test case name who's CAN Message user want.
        :return: CAN Message corresponding to given test case name
        """
        try:
            if test_description in self.get_test_case_list():
                excel_dict = self._get_excel_output_dict()
                input_message = excel_dict[test_description][ParseInputExcel._INPUT_MESSAGE]
                return json.loads(input_message)
            else:
                robot_print_error("The given Test case name is not match in input excel file...!!!",
                                  print_in_report=True)
        except json.JSONDecodeError as json_err:
            robot_print_error(f"Error to decode the publish payload, EXCEPTION: {json_err}", print_in_report=True)
        except IOError as ioerr:
            robot_print_error(f"The given Test case name is not match in InputExcel...!!!"
                              f"\nPlease check the test case name...!!!, EXCEPTION: {ioerr}", print_in_report=True)

    def get_expected_output(self, test_description) -> json:
        """
        This method provide the Expected Output form the excel file corresponding to the give test case name
        :param test_description: test case name who's Expected Output user want.
        :return: Dictionary of Expected Output corresponding to given test case name
        """
        expect_output = {}
        try:
            if test_description in self.get_test_case_list():
                excel_dict = self._get_excel_output_dict()
                expect_output = excel_dict[test_description][ParseInputExcel._EXPECTED_OUTPUT]
                return json.loads(expect_output)
            else:
                robot_print_error("The given Test case name is not match in input excel file...!!!",
                                  print_in_report=True)
        except json.JSONDecodeError as jsonerr:
            robot_print_error(f"There is an error to decode the excepted Output, EXCEPTION: {jsonerr}",
                              print_in_report=True)
            return expect_output
        except IOError as ioerr:
            robot_print_error(f"The given Test case name is not match in InputExcel...!!!"
                              f"\nPlease check the test case name...!!!\nEXCEPTION: {ioerr}",
                              print_in_report=True)

    def get_kpi_value(self, test_description) -> str:
        """
        This method provide the KPI value form the excel file corresponding to the give test case name
        :param test_description: test case name who's KPI value user want.
        :return: KPI value corresponding to given test case name
        """
        try:
            if test_description in self.get_test_case_list():
                excel_dict = self._get_excel_output_dict()
                return excel_dict[test_description][ParseInputExcel._KIP_VALUE]
            else:
                robot_print_error("The given Test case name is not match in input excel file...!!!",
                                  print_in_report=True)
        except IOError as ioerr:
            robot_print_error(f"The given Test case name is not match in InputExcel...!!!"
                              f"\nPlease check the test case name...!!!\nEXCEPTION: {ioerr}",
                              print_in_report=True)

    def set_test_output(self, test_description, value):
        """
        Write the Test Output into the excel file corresponding to give test case
        :param test_description: test case name who's Test output user want to set
        :param value: Test output comes form the Broker response
        """
        try:
            book = load_workbook(self.input_file_path)
            sheet = book.active
            test_case_index = self.input_sheet[self.input_sheet["Test_Step_Description"] == test_description].index
            index = int(test_case_index[0])
            sheet.cell(row=index + 2, column=6).value = value
            book.save(self.input_file_path)
        except IndexError as index_exception:
            robot_print_error(f"Error to set the test case output, EXCEPTION:{index_exception}", print_in_report=True)
        except ValueError as val_err:
            robot_print_error(f"Error to set the output inside excel, EXCEPTION: {val_err}", print_in_report=True)

    def set_test_result_status(self, test_description, status):
        """
        Write the Test Status (PASS/FAILED) into the excel file corresponding to give test case
        :param test_description: test case name who's Status user want to set
        :param status: status of test case depends upon the Comparision of Expected value and Broker Response
        """
        try:
            book = load_workbook(self.input_file_path)
            sheet = book.active
            test_case_index = self.input_sheet[self.input_sheet["Test_Step_Description"] == test_description].index
            index = int(test_case_index[0])
            sheet.cell(row=index + 2, column=7).value = status
            book.save(self.input_file_path)
        except IndexError as index_exception:
            robot_print_error(f"Error to set the result status, EXCEPTION: {index_exception}",
                              print_in_report=True)
        except ValueError as val_err:
            robot_print_error(f"Error to set the result status into excel, EXCEPTION: {val_err}")
