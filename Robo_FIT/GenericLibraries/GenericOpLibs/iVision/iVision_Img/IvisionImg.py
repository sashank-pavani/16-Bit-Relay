from Robo_FIT.GenericLibraries.GenericOpLibs.iVision.iVision_Img.ConfigurationManager import ConfigurationManager
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import *
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass
import os
import time
import pandas as pd
import numpy as np
from datetime import timedelta
import openpyxl
from openpyxl import load_workbook
import cv2
import shutil
import glob
import re

class IvisionImg:
    def __init__(self):
        """Initialize with the server URL and load expected colors."""
        self.config_manager = ConfigurationManager()
        self.common_keyword = CommonKeywordsClass()
        self.test_case_name = None
        self.excel_file_path_for_pattern_matching = self.config_manager.get_path('excel_file_path')
        self.images_directory_for_pattern_matching = self.config_manager.get_path('images_directory')
        self.excel_file_path = self.config_manager.get_blinking_parameters("excel_path")
        self.images_directory = self.config_manager.get_blinking_parameters("buffer_path")
        self.threshold = self.config_manager.get_threshold()
        self.black_bg = str(self.config_manager.get_black_bg()).lower() == "true"

        self.initial_state = ""
        self.current_state = ""
        self.start_flag = 0
        self.end_flag = 0
        self.full_cycle_end = 0

    def set_ivision_test_case_folder(self, test_case_name: str):
        """Sets the test case name for the instance.

        Args:
            test_case_name (str): The name of the test case.
        """

        try:
            self.test_case_name = test_case_name
            robot_print_info(f"Test case name set to: {self.test_case_name}")
        except Exception as e:
            robot_print_error(f"Error setting test case name: {str(e)}")

    def ivision_input_image(self, imag_path):
        """Save input image to the test report path based on the test case name."""
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")

            input_image_dir = self.get_ivision_image_path()
            # Save the input image as 'input_image.png' instead of its original name
            input_image_path = os.path.join(input_image_dir, "input_image.png")

            with open(imag_path, 'rb') as src:
                with open(input_image_path, 'wb') as dst:
                    dst.write(src.read())

            robot_print_info(f"Input image saved at: {input_image_path}")
            return input_image_path
        except Exception as e:
            robot_print_error(f"Error saving input image: {e}")
            raise RuntimeError(f"Error in iVision_input_image: {e}")

    def get_ivision_image_path(self):
        """Retrieves the path to the iVision image directory.

        Returns:
            str: The path to the iVision image directory.

        Raises:
            ValueError: If test_case_name is not set.
            OSError: If directory creation fails.
        """
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")

            subfolders = [self.config_manager.get_ivision_base_folder(),
                          self.config_manager.get_ivision_image_subfolder(), self.test_case_name]
            image_path = os.path.join(self.common_keyword.get_report_path(), *subfolders)

            if not os.path.isdir(image_path):
                os.makedirs(image_path, mode=0o777)
                time.sleep(1)
            return image_path
        except OSError as os_err:
            robot_print_error(f"Error creating the IMAGE directory: {os_err}")
            raise
        except Exception as e:
            robot_print_error(f"Unexpected error in get_ivision_image_path: {str(e)}")
            raise

    def get_ivision_result_path(self):
        """Retrieves the path to the iVision result directory.

        Returns:
            str: The path to the iVision result directory.

        Raises:
            ValueError: If test_case_name is not set.
            OSError: If directory creation fails.
        """
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")

            subfolders = [self.config_manager.get_ivision_base_folder(),
                          self.config_manager.get_ivision_result_subfolder(), self.test_case_name]
            result_path = os.path.join(self.common_keyword.get_report_path(), *subfolders)

            if not os.path.isdir(result_path):
                os.makedirs(result_path, mode=0o777)
                time.sleep(1)
            return result_path
        except OSError as os_err:
            robot_print_error(f"Error creating the RESULT directory: {os_err}")
            raise
        except Exception as e:
            robot_print_error(f"Unexpected error in get_ivision_result_path: {str(e)}")
            raise

    def ivision_copy_img(self, screenshot_path):
        """
        Robustly copy the specific image transferred by transfer_specific_image_rbf.
        Args:
            screenshot_path (str): Path to image or directory containing images.
        """
        try:
            if not hasattr(self, 'test_case_name') or not self.test_case_name:
                raise ValueError("Test case name is not set.")

            report_image_dir = os.path.join(self.get_ivision_image_path())

            if os.path.isfile(screenshot_path):

                specific_image_path = screenshot_path
            else:
                image_pattern = os.path.join(screenshot_path, "*.jpg")
                image_files = glob.glob(image_pattern)

                if not image_files:
                    raise FileNotFoundError(f"No image files found in: {screenshot_path}")

                specific_image_path = max(image_files, key=os.path.getmtime)

            image_filename = os.path.basename(specific_image_path)

            specific_image_report_path = os.path.join(report_image_dir, image_filename)

            shutil.copyfile(specific_image_path, specific_image_report_path)

            robot_print_info(f"Specific image copied from {specific_image_path} to: {specific_image_report_path}")
            return specific_image_report_path

        except Exception as e:
            robot_print_error(f"Error in ivision_copy_img: {e}")
            raise RuntimeError(f"Failed to handle specific image: {e}")

    def ivision_copy_specific_img(self, screenshot_path):
        """
        Copy the specific image transferred and renamed by the move image function.
        Args:
            screenshot_path (str): Full path to the image file transferred by the move image function.
        """
        try:

            if not hasattr(self, 'test_case_name') or not self.test_case_name:
                raise ValueError("Test case name is not set.")

            if not os.path.isfile(screenshot_path):
                raise FileNotFoundError(f"Image file not found at path: {screenshot_path}")

            report_image_dir = os.path.join(self.get_ivision_image_path())
            os.makedirs(report_image_dir, exist_ok=True)

            image_filename = os.path.basename(screenshot_path)

            specific_image_report_path = os.path.join(report_image_dir, image_filename)
            shutil.copyfile(screenshot_path, specific_image_report_path)

            robot_print_info(f"Specific image copied from {screenshot_path} to: {specific_image_report_path}")
            return specific_image_report_path

        except Exception as e:
            robot_print_error(f"Error in ivision_copy_img: {e}")
            raise RuntimeError(f"Failed to handle specific image: {e}")

    def ivision_copy_all_files(self, return_files):
        """
        Copy all files from the destination folder to the test report path.

        Args:
            return_files (str): Either the path to files or a success message containing the path.

        Returns:
            str: Success message with the path where files have been copied.
        """
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")
            if "Successfully copied" in return_files:
                import re
                path_match = re.search(r'to (.*)\.$', return_files)
                if path_match:
                    return_files = path_match.group(1).strip()
                else:
                    raise ValueError("Could not extract path from the success message")

            if not os.path.exists(return_files):
                raise FileNotFoundError(f"Source directory does not exist: {return_files}")

            test_report_path = self.get_ivision_image_path()
            ivision_folder_path = os.path.join(test_report_path)

            os.makedirs(ivision_folder_path, exist_ok=True)

            robot_print_info(f"Copying files to: {ivision_folder_path} from: {return_files}")
            files = []
            for item in os.listdir(return_files):
                item_path = os.path.join(return_files, item)
                if os.path.isfile(item_path):
                    files.append(item)

            for file in files:
                src_path = os.path.join(return_files, file)
                dst_path = os.path.join(ivision_folder_path, file)
                shutil.copy2(src_path, dst_path)

            robot_print_info(f"Successfully copied {len(files)} files to {ivision_folder_path}.")
            return f"Successfully copied {len(files)} files to iVision folder at {ivision_folder_path}."

        except Exception as e:
            error_message = f"Error in ivision_copy_all_files: {str(e)}"
            robot_print_error(error_message)
            robot_print_error(f"Source path attempted: {return_files}")

            raise RuntimeError(error_message)

    def ivision_auto_check_pattern_matching(self, test_case_name, object_names, threshold=None):
        """
        Performs pattern matching on an input image using multiple object templates and provides a summary.

        Args:
            test_case_name (str): The name of the test case to set the directory and context.
            object_names (list or str): List of object names or a string representation of list to match.
            threshold (float, optional): Similarity threshold for template matching. Defaults to a class attribute.

        Returns:
            tuple: A tuple containing:
                - all_match (bool): True if all objects matched above threshold, False otherwise.
                - result_summary (dict): Dictionary with match details for each object, including confidence and match status.

        Raises:
            Exception: Propagates exceptions encountered during the process.
        """
        try:
            if threshold is None:
                threshold = self.threshold

            threshold = float(threshold)

            if isinstance(object_names, str):
                object_names = eval(object_names)

            if len(object_names) > 10:
                raise ValueError("You can only pass up to 10 object names.")

            self.set_ivision_test_case_folder(test_case_name)
            image_path = self.get_ivision_image_path()
            robot_print_info(f'Input image taken from: {image_path}')

            input_images = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            if not input_images:
                raise FileNotFoundError(f"No input image found in {image_path}")
            input_image = input_images[0]
            image_name = os.path.join(image_path, input_image)

            large_image = cv2.imread(image_name)
            if large_image is None:
                raise ValueError("Could not load large image")

            df = pd.read_excel(self.excel_file_path, engine='openpyxl')
            second_row = df.iloc[1]
            roi_col = second_row['Unnamed: 1']
            result_summary = {}

            for object_name in object_names:
                try:
                    row = df[df.iloc[:, 3] == object_name]
                    if row.empty:
                        raise ValueError(f"Object name '{object_name}' not found in the Excel sheet.")

                    template_image_name = str(row.iloc[0, 2])
                    robot_print_info(f'Template image name: {template_image_name}')
                    template_image_path = os.path.join(self.images_directory, template_image_name)
                    robot_print_info(f'Template image path: {template_image_path}')

                    if not os.path.isfile(template_image_path):
                        raise FileNotFoundError(
                            f"Template image '{template_image_name}' not found in {self.images_directory}")

                    template = cv2.imread(template_image_path, cv2.IMREAD_UNCHANGED)
                    if template is None:
                        raise ValueError(f"Could not load template image: {template_image_path}")

                    result = self._match_template_on_image(large_image, template, threshold)
                    if result:
                        robot_print_info(f"Template '{object_name}' matched at {result['top_left']} "
                                         f"with confidence {result['confidence']:.2f}, scale {result['scale']:.2f}, "
                                         f"avg color {result['avg_color_rgb']} ({result['avg_color_hex']})")
                        result_summary[object_name] = {
                            "confidence": round(result["confidence"], 4),
                            "matched": True,
                            "top_left": result["top_left"],
                            "bottom_right": result["bottom_right"]
                        }
                    else:
                        robot_print_error(f"Template '{object_name}' did not match above threshold.")
                        result_summary[object_name] = {
                            "confidence": 0.0,
                            "matched": False
                        }

                except Exception as e:
                    robot_print_error(f"Failed to process object '{object_name}': {e}")
                    result_summary[object_name] = {
                        "confidence": 0.0,
                        "matched": False
                    }

            if self.black_bg:
                black_image = np.zeros_like(large_image)
                for result in result_summary.values():
                    if result["matched"]:
                        top_left = result["top_left"]
                        bottom_right = result["bottom_right"]
                        black_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = \
                            large_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                result_image = black_image
            else:
                for object_name in object_names:
                    result = result_summary.get(object_name)
                    if result and result["matched"]:
                        top_left = result["top_left"]
                        bottom_right = result["bottom_right"]
                        cv2.rectangle(large_image, top_left, bottom_right, (0, 255, 0), 2)
                result_image = large_image

            result_path = os.path.join(self.get_ivision_result_path(), f'auto_pattern_matching_{object_names}.png')
            cv2.imwrite(result_path, result_image)
            robot_print_info(f"Result image for '{object_names}' saved at: {result_path}")

            all_match = all(result["matched"] for result in result_summary.values())
            return all_match, result_summary

        except Exception as e:
            robot_print_error(f"Template matching failed: {e}")
            return {}

    def ivision_check_pattern_matching(self, test_case_name, template_image_path, threshold=None):
        """
        Performs pattern matching on an input image using a single template image.

        Args:
            test_case_name (str): The name of the test case to set the directory and context.
            template_image_path (str): The file path to the template image.
            threshold (float, optional): Similarity threshold for template matching. Defaults to a class attribute.

        Returns:
            bool: True if the template match is above threshold, False otherwise.

        Raises:
            Exception: Propagates exceptions encountered during the process.
        """
        try:
            if threshold is None:
                threshold = self.threshold

            threshold = float(threshold)

            if not os.path.isfile(template_image_path):
                raise FileNotFoundError(f"Template image not found at: {template_image_path}")

            self.set_ivision_test_case_folder(test_case_name)
            image_path = self.get_ivision_image_path()
            robot_print_info(f'Input image taken from: {image_path}')

            input_images = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            if not input_images:
                raise FileNotFoundError(f"No input image found in {image_path}")
            input_image = input_images[0]
            image_name = os.path.join(image_path, input_image)

            large_image = cv2.imread(image_name)
            if large_image is None:
                raise ValueError("Could not load large image")

            template = cv2.imread(template_image_path, cv2.IMREAD_UNCHANGED)
            if template is None:
                raise ValueError(f"Could not load template image: {template_image_path}")

            result = self._match_template_on_image(large_image, template, threshold)
            if result:
                robot_print_info(f"Template matched at {result['top_left']} "
                                 f"with confidence {result['confidence']:.2f}, scale {result['scale']:.2f}")

                if self.black_bg:
                    black_image = np.zeros_like(large_image)
                    top_left = result["top_left"]
                    bottom_right = result["bottom_right"]
                    black_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = \
                        large_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                    result_image = black_image
                else:
                    cv2.rectangle(large_image, result["top_left"], result["bottom_right"], (0, 255, 0), 2)
                    result_image = large_image

                result_path = os.path.join(self.get_ivision_result_path(),
                                           f'single_pattern_matching_{test_case_name}.png')
                cv2.imwrite(result_path, result_image)
                robot_print_info(f"Result image saved at: {result_path}")
                return True
            else:
                robot_print_warning("Template did not match above threshold.")
                return False

        except Exception as e:
            robot_print_error(f"Template matching failed: {e}")
            return False

    def _rgb_to_hex(self, rgb):
        """
        Converts an RGB color tuple to its corresponding hexadecimal color string.

        Returns:
            str: The hexadecimal color string in format '#rrggbb'.
        """
        return '#{:02x}{:02x}{:02x}'.format(int(rgb[0]), int(rgb[1]), int(rgb[2]))

    def _detect_template_roi(self, template):
        """
        Detects and returns the region of interest (ROI) in a template image.

        Returns:
            tuple or None: A tuple (x, y, w, h) defining the top-left corner and size of the ROI,
                         or None if no suitable ROI is found.
        """
        try:
            if template.shape[2] == 4:
                try:
                    alpha = template[:, :, 3]
                    non_transparent = cv2.findNonZero(alpha)
                    if non_transparent is not None:
                        x, y, w, h = cv2.boundingRect(non_transparent)
                        if w > 0 and h > 0:
                            return (x, y, w, h)
                except Exception as e:
                    robot_print_error(f"[WARNING] Failed to detect ROI from alpha channel: {e}")

            try:
                gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
                edges = cv2.Canny(gray, 50, 150)
                non_zero = cv2.findNonZero(edges)
                if non_zero is not None:
                    x, y, w, h = cv2.boundingRect(non_zero)
                    if w > 0 and h > 0:
                        pad = 2
                        x = max(0, x - pad)
                        y = max(0, y - pad)
                        w = min(template.shape[1] - x, w + 2 * pad)
                        h = min(template.shape[0] - y, h + 2 * pad)
                        return (x, y, w, h)
            except Exception as e:
                robot_print_error(f"Failed to detect ROI using edge detection: {e}")

        except Exception as e:
            robot_print_error(f"Unexpected error in _detect_template_roi: {e}")

        return None

    def _match_template_on_image(self, large_image, template, threshold=0.8, min_template_size=10):
        """
        Performs template matching on a large image to locate the template.

        Returns:
            dict: A dictionary with match details:
                - 'top_left': (x, y) coordinates of the match's top-left corner.
                - 'bottom_right': (x, y) coordinates of the match's bottom-right corner.
                - 'confidence': Similarity score of the match.
                - 'scale': Scaling factor that yielded the best match.
                - 'avg_color_rgb': Average RGB color of the matched region.
                - 'avg_color_hex': Average color in hexadecimal format.
            or
            bool: False if no match is found above the threshold or an error occurs.
        """
        try:
            roi = self._detect_template_roi(template)
            if roi is not None:
                x, y, w, h = roi
                template = template[y:y + h, x:x + w]

            try:
                large_gray = cv2.cvtColor(large_image, cv2.COLOR_BGR2GRAY)
                template_gray = cv2.cvtColor(template,
                                             cv2.COLOR_BGR2GRAY if template.shape[2] == 3 else cv2.COLOR_BGRA2GRAY)
            except Exception as e:
                robot_print_error(f" Failed to convert images to grayscale: {e}")
                return None

            large_gray = cv2.GaussianBlur(large_gray, (3, 3), 0)
            template_gray = cv2.GaussianBlur(template_gray, (3, 3), 0)

            template_h, template_w = template_gray.shape
            large_h, large_w = large_gray.shape

            min_scale = min_template_size / min(template_w, template_h)
            max_scale = min(large_w / template_w, large_h / template_h) * 0.9
            scales = np.linspace(max(0.1, min_scale), min(max_scale, 5.0), 30)

            best_val = -1
            best_loc = None
            best_scale = 1.0

            for scale in scales:
                try:
                    scaled_template = cv2.resize(template_gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
                    scaled_h, scaled_w = scaled_template.shape

                    if scaled_h > large_h or scaled_w > large_w:
                        continue

                    result = cv2.matchTemplate(large_gray, scaled_template, cv2.TM_CCOEFF_NORMED)
                    _, max_val, _, max_loc = cv2.minMaxLoc(result)

                    if max_val > best_val:
                        best_val = max_val
                        best_loc = max_loc
                        best_scale = scale
                except Exception as e:
                    robot_print_warning(f"Failed at scale {scale:.2f}: {e}")
                    continue

            if best_val >= threshold:
                scaled_template_h = int(template_h * best_scale)
                scaled_template_w = int(template_w * best_scale)
                top_left = best_loc
                bottom_right = (top_left[0] + scaled_template_w, top_left[1] + scaled_template_h)

                try:
                    matched_region = large_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                    avg_color_bgr = np.mean(matched_region, axis=(0, 1))
                    avg_color_rgb = avg_color_bgr[::-1]
                    avg_color_rgb = tuple(avg_color_rgb)
                    avg_color_hex = self._rgb_to_hex(avg_color_rgb)
                except Exception as e:
                    robot_print_error(f"Failed to compute average color: {e}")
                    avg_color_rgb = (0, 0, 0)
                    avg_color_hex = "#000000"

                return {
                    "top_left": top_left,
                    "bottom_right": bottom_right,
                    "confidence": best_val,
                    "scale": best_scale,
                    "avg_color_rgb": avg_color_rgb,
                    "avg_color_hex": avg_color_hex
                }
            else:
                robot_print_debug(f"No match found above threshold ({threshold})")
                return False

        except Exception as e:
            robot_print_error(f"[ERROR] Template matching failed: {e}")
            return False

    def ivision_auto_check_pattern_matching_with_roi(self, test_case_name, object_names, threshold=None):
        """
        Performs template matching on a large image to locate the template.

        Returns:
            dict: A dictionary with match details:
                - 'top_left': (x, y) coordinates of the match's top-left corner.
                - 'bottom_right': (x, y) coordinates of the match's bottom-right corner.
                - 'confidence': Similarity score of the match.
                - 'scale': Scaling factor that yielded the best match.
                - 'avg_color_rgb': Average RGB color of the matched region.
                - 'avg_color_hex': Average color in hexadecimal format.
            or
            bool: False if no match is found above the threshold or an error occurs.
        """
        try:
            if threshold is None:
                threshold = self.threshold
            threshold = float(threshold)

            if isinstance(object_names, str):
                object_names = eval(object_names)

            if len(object_names) > 10:
                raise ValueError("You can only pass up to 10 object names.")

            self.set_ivision_test_case_folder(test_case_name)
            image_path = self.get_ivision_image_path()
            robot_print_info(f'Input image taken from: {image_path}')

            input_images = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            if not input_images:
                raise FileNotFoundError(f"No input image found in {image_path}")
            input_image = input_images[0]
            image_name = os.path.join(image_path, input_image)

            original_image = cv2.imread(image_name)
            if original_image is None:
                raise ValueError("Could not load large image")

            df = pd.read_excel(self.excel_file_path_for_pattern_matching, engine='openpyxl')
            second_row = df.iloc[1]
            roi_col = second_row['Unnamed: 1']

            result_summary = {}

            for object_name in object_names:
                try:
                    row = df[df.iloc[:, 3] == object_name]
                    if row.empty:
                        raise ValueError(f"Object name '{object_name}' not found in the Excel sheet.")

                    roi = row.iloc[0, 1]
                    if pd.isna(roi) or str(roi).strip() == "":
                        raise ValueError(f"ROI not provided for object '{object_name}' in Excel sheet.")

                    roi_str = str(roi)
                    if ',' in roi_str:
                        formatted_roi = ' '.join([x.strip() for x in roi_str.split(',')])
                    else:
                        groups = re.findall(r'\d{1,4}', roi_str)
                        formatted_roi = " ".join(groups)

                    roi_coords = list(map(int, formatted_roi.split()))
                    if len(roi_coords) != 4:
                        raise ValueError(f"Invalid ROI format for object '{object_name}': {roi_coords}")

                    x_min, y_min, width, height = roi_coords
                    robot_print_info(f'{roi_coords}')
                    x_max = x_min + width
                    y_max = y_min + height
                    roi_offset = (x_min, y_min)
                    robot_print_info(f'{x_max},{y_max}')

                    robot_print_info(f'ROI: x_min={x_min}, y_min={y_min}, x_max={x_max}, y_max={y_max}')
                    roi_image = original_image[y_min:y_max, x_min:x_max]

                    template_image_name = str(row.iloc[0, 2])
                    template_image_path = os.path.join(self.images_directory_for_pattern_matching, template_image_name)
                    robot_print_info(f'Template image path: {template_image_path}')

                    if not os.path.isfile(template_image_path):
                        raise FileNotFoundError(f"Template image '{template_image_name}' not found.")

                    template = cv2.imread(template_image_path, cv2.IMREAD_UNCHANGED)
                    if template is None:
                        raise ValueError(f"Could not load template image: {template_image_path}")

                    result = self._match_template_on_image(roi_image, template, threshold)
                    if result:
                        top_left = (result["top_left"][0] + x_min, result["top_left"][1] + y_min)
                        bottom_right = (result["bottom_right"][0] + x_min, result["bottom_right"][1] + y_min)
                        robot_print_info(f"Template '{object_name}' matched at {top_left} "
                                         f"with confidence {result['confidence']:.2f}, scale {result['scale']:.2f}, "
                                         f"avg color {result['avg_color_rgb']} ({result['avg_color_hex']})")
                        result_summary[object_name] = {
                            "confidence": round(result["confidence"], 4),
                            "matched": True,
                            "top_left": top_left,
                            "bottom_right": bottom_right
                        }

                        # Save result image for matched object
                        result_image = np.zeros_like(original_image) if self.black_bg else original_image.copy()
                        if self.black_bg:
                            result_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = \
                                original_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                        else:
                            cv2.rectangle(result_image, top_left, bottom_right, (0, 255, 0), 2)

                        result_path = os.path.join(self.get_ivision_result_path(),
                                                   f'auto_pattern_matching_{object_name}.png')
                        cv2.imwrite(result_path, result_image)
                        robot_print_info(f"Result image for '{object_name}' saved at: {result_path}")

                    else:
                        robot_print_error(f"Template '{object_name}' did not match above threshold.")
                        result_summary[object_name] = {
                            "confidence": 0.0,
                            "matched": False
                        }

                except Exception as e:
                    robot_print_error(f"Failed to process object '{object_name}': {e}")
                    result_summary[object_name] = {
                        "confidence": 0.0,
                        "matched": False
                    }

            all_match = all(result["matched"] for result in result_summary.values())
            return all_match, result_summary

        except Exception as e:
            robot_print_error(f"Template matching failed: {e}")
            return False, {}
       

    def ivision_check_pattern_matching_with_roi(self, test_case_name, template_image_path, roi=None, threshold=None):
        """
        Performs pattern matching on an input image using a single template image, within a specified ROI.

        Args:
            test_case_name (str): The name of the test case to set the directory and context.
            template_image_path (str): The file path to the template image.
            threshold (float, optional): Similarity threshold for template matching. Defaults to a class attribute.
            roi (str or tuple): Region of interest in the format "x_min, y_min, width, height".

        Returns:
            bool: True if the template match is above threshold, False otherwise.
        """
        try:
            if threshold is None:
                threshold = self.threshold
            threshold = float(threshold)

            if not os.path.isfile(template_image_path):
                raise FileNotFoundError(f"Template image not found at: {template_image_path}")

            if not roi:
                raise ValueError("ROI not provided by user.")

            if isinstance(roi, str):
                roi = list(map(int, roi.split(',')))
            if len(roi) != 4:
                raise ValueError("ROI must be in the format 'x_min, y_min, width, height'")

            x_min, y_min, width, height = roi
            x_max = x_min + width
            y_max = y_min + height
            roi_offset = (x_min, y_min)

            self.set_ivision_test_case_folder(test_case_name)
            image_path = self.get_ivision_image_path()
            robot_print_info(f'Input image taken from: {image_path}')

            input_images = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            if not input_images:
                raise FileNotFoundError(f"No input image found in {image_path}")
            input_image = input_images[0]
            image_name = os.path.join(image_path, input_image)

            original_image = cv2.imread(image_name)
            if original_image is None:
                raise ValueError("Could not load large image")

            roi_image = original_image[y_min:y_max, x_min:x_max]

            template = cv2.imread(template_image_path, cv2.IMREAD_UNCHANGED)
            if template is None:
                raise ValueError(f"Could not load template image: {template_image_path}")

            result = self._match_template_on_image(roi_image, template, threshold)
            if result:
                top_left = (result["top_left"][0] + x_min, result["top_left"][1] + y_min)
                bottom_right = (result["bottom_right"][0] + x_min, result["bottom_right"][1] + y_min)

                robot_print_info(f"Template matched at {top_left} "
                                 f"with confidence {result['confidence']:.2f}, scale {result['scale']:.2f}")

                if self.black_bg:
                    black_image = np.zeros_like(original_image)
                    black_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = \
                        original_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                    result_image = black_image
                else:
                    result_image = original_image.copy()
                    cv2.rectangle(result_image, top_left, bottom_right, (0, 255, 0), 2)

                result_path = os.path.join(self.get_ivision_result_path(),
                                           f'single_pattern_matching_{test_case_name}.png')
                cv2.imwrite(result_path, result_image)
                robot_print_info(f"Result image saved at: {result_path}")
                return True
            else:
                robot_print_warning("Template did not match above threshold.")
                return False
        except Exception as e:
            robot_print_error(f"Template matching failed: {e}")
            return False, {}

    def ivision_check_pattern_matching_for_blinking(self, test_case_name, input_image_path, template_image_path,
                                                    threshold=None,
                                                    save_image=False, save_path=None):
        """
        This function is used for pattern matching when the blink measurement function is used.
        """
        try:
            if threshold is None:
                threshold = self.threshold

            threshold = float(threshold)

            if not os.path.isfile(template_image_path):
                raise FileNotFoundError(f"Template image not found at: {template_image_path}")

            self.set_ivision_test_case_folder(test_case_name)

            if not os.path.isfile(input_image_path):
                raise FileNotFoundError(f"Input image not found at: {input_image_path}")

            robot_print_info(f'Input image taken from: {input_image_path}')

            large_image = cv2.imread(input_image_path)
            if large_image is None:
                raise ValueError("Could not load input image")

            template = cv2.imread(template_image_path, cv2.IMREAD_UNCHANGED)
            if template is None:
                raise ValueError(f"Could not load template image: {template_image_path}")

            result = self._match_template_on_image(large_image, template, threshold)
            if result:
                robot_print_info(f"Template matched at {result['top_left']} "
                                 f"with confidence {result['confidence']:.2f}, scale {result['scale']:.2f}")

                if self.black_bg:
                    black_image = np.zeros_like(large_image)
                    top_left = result["top_left"]
                    bottom_right = result["bottom_right"]
                    black_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]] = \
                        large_image[top_left[1]:bottom_right[1], top_left[0]:bottom_right[0]]
                    result_image = black_image
                else:
                    cv2.rectangle(large_image, result["top_left"], result["bottom_right"], (0, 255, 0), 2)
                    result_image = large_image

                if save_image:
                    if save_path:
                        result_path = save_path
                    else:
                        result_path = os.path.join(self.get_ivision_result_path(),
                                                   f'single_pattern_matching_{test_case_name}.png')
                    cv2.imwrite(result_path, result_image)
                    # robot_print_info(f"Result image saved at: {result_path}")

                return True
            else:
                return False

        except Exception as e:
            robot_print_error(f"Template matching failed: {e}")
            return False

    def resize_images_in_path(self, buffer_path: str, width: int, height: int):
        """
        Resize all image files in the specified directory to the given width and height.

        Supports filenames with timestamps in:
        - Format 1: image_YYYYMMDD_HH_MM_SS_micro.jpg
        - Format 2: -1-07d_00h_54m_14.046-img-1-

        Parameters:
            buffer_path (str): Path to the folder containing images to resize.
            width (int): Target width for the resized images.
            height (int): Target height for the resized images.

        Returns:
            bool: True if at least one image was processed, False if folder is empty or an error occurred.
        """
        try:
            robot_print_info(f"Resizing images in path: {buffer_path}")

            if not os.path.exists(buffer_path):
                robot_print_info(f"Provided path does not exist: {buffer_path}")
                return False

            image_files = [f for f in os.listdir(buffer_path)
                           if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff'))]

            if not image_files:
                robot_print_info(f"No image files found in: {buffer_path}")
                return False

            for filename in image_files:
                image_path = os.path.join(buffer_path, filename)
                img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)

                # Extract timestamp string from filename
                timestamp_str = "unknown"
                if "image_" in filename:
                    try:
                        timestamp_str = filename.split("image_")[1].split(".jpg")[0]
                    except Exception:
                        pass
                else:
                    match = re.search(r"(\d{2})d_(\d{2})h_(\d{2})m_(\d{2})\.(\d+)", filename)
                    if match:
                        days, hours, minutes, seconds, milliseconds = match.groups()
                        timestamp_str = f"{days}d_{hours}h_{minutes}m_{seconds}_{milliseconds}"

                if img is not None:
                    if img.shape[1] != width or img.shape[0] != height:
                        resized_img = cv2.resize(img, (width, height))
                        cv2.imwrite(image_path, resized_img)
                        robot_print_info(f"Resized '{filename}' ({timestamp_str}) to {width}x{height}")
                        robot_print_info(f"Resized image saved at: {image_path}")
                    else:
                        robot_print_info(f"'{filename}' ({timestamp_str}) already {width}x{height}, skipping")
                else:
                    robot_print_info(f"Failed to read image '{filename}'")

            return True

        except Exception as e:
            robot_print_info(f"Exception during image resizing: {e}")
            return False

    def capture_images_with_timestamp(self, buffer_limit=300):
        """
        Capture images from the webcam and save them with timestamped filenames.

        This function captures frames from the default webcam and saves them as image files
        in the buffer path specified in the configuration. Each image is named using the
        current timestamp to ensure uniqueness.

        Parameters:
            buffer_limit (int): Maximum number of images to capture. Default is 300.
        """

        save_dir = self.config_manager.get_blinking_parameters('buffer_path')
        robot_print_info(f"Execution started :")
        robot_print_info(f"save_dir ==== :{save_dir}")
        self.delete_all_files_in_folder(save_dir)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            print("Error: Could not open webcam.")
            return
        print(f"Capturing images... (up to {buffer_limit} images)")
        image_count = 0

        while image_count < buffer_limit:
            ret, frame = cap.read()

            if not ret:
                print("Error: Could not read frame.")
                break
            timestamp = datetime.now().strftime("%Y%m%d_%H_%M_%S_%f")
            file_name = f"{save_dir}/image_{timestamp}.jpg"
            cv2.imwrite(file_name, frame)
            print(f"Saved {file_name}")
            image_count += 1
        cap.release()
        cv2.destroyAllWindows()
        print(f"Done! {image_count} images saved in '{save_dir}'.")

    def list_files_in_folder(self):
        """
        List all files in the buffer path specified in the configuration.

        Returns:
            list: A list of filenames present in the buffer folder.
        """

        # get_destination_path = self.config_manager.get_buffer_path()
        get_destination_path = self.config_manager.get_blinking_parameters('buffer_path')
        try:
            files = os.listdir(get_destination_path)

            files = [f for f in files if os.path.isfile(os.path.join(get_destination_path, f))]

            return files
        except Exception as e:
            return str(e)

    def save_files_to_excel(self):
        """
        Save the list of image filenames from the buffer folder into an Excel file.

        The Excel file path is taken from the configuration. If the file exists, it will be overwritten.

        Returns:
            str: Success message or error message.
        """

        files = self.list_files_in_folder()
        # get_excel_path = self.config_manager.get_excel_path()
        get_excel_path = self.config_manager.get_blinking_parameters('excel_path')
        try:

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            for index, file_name in enumerate(files, start=1):
                sheet[f"A{index}"] = file_name

            workbook.save(get_excel_path)
            return f"Excel file created and saved at {get_excel_path}"
        except Exception as e:
            return str(e)

    def read_column_from_excel(self, sheet_name, column_letter):
        """
        Read all values from a specific column in a given Excel sheet.

        Args:
            sheet_name (str): Name of the sheet to read from.
            column_letter (str): Excel column letter (e.g., 'A', 'B').

        Returns:
            list: Values from the specified column or error message.
        """

        # get_excel_path = self.config_manager.get_update_file_path()
        get_excel_path = self.config_manager.get_blinking_parameters('update_file_path')
        try:

            workbook = openpyxl.load_workbook(get_excel_path)

            sheet = workbook[sheet_name]

            column_values = [cell.value for cell in sheet[column_letter]]

            return column_values
        except Exception as e:
            return str(e)

    def delete_all_files_in_folder(self, folder_path):
        """
        Delete all files inside the specified folder.

        Args:
            folder_path (str): Path to the folder whose contents should be deleted.
        """

        try:
            if not os.path.exists(folder_path):
                robot_print_info(f"Error: The folder '{folder_path}' does not exist.")
                return

            files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

            for file in files:
                file_path = os.path.join(folder_path, file)
                os.remove(file_path)
            robot_print_info(f"All files deleted successfully.")

        except Exception as e:
            robot_print_info(f"Exception occured == {e}")

    def format_inside_excel(self):
        """
        Format and structure data inside the Excel workbook specified in the config.

        This includes:
        - Copying filenames to a new sheet.
        - Splitting filenames into components.
        - Creating additional structured columns.
        - Saving the formatted data to a new Excel file.
        """

        try:
            self.save_files_to_excel()
            self.create_new_sheet()

            file_path = self.config_manager.get_blinking_parameters('excel_path')
            robot_print_info(f"Excel path: {file_path}")
            wb = load_workbook(file_path)
            ws = wb.active

            for row in ws.iter_rows(min_col=1, max_col=1, min_row=1):
                for cell in row:
                    ws.cell(row=cell.row, column=2, value=cell.value)

            df = pd.DataFrame(ws.values)
            robot_print_info(f"Initial DataFrame shape: {df.shape}")

            if 0 in df.columns:
                df_expanded = df[0].astype(str).str.split('_', expand=True)
                df[list(range(df.shape[1], df.shape[1] + df_expanded.shape[1]))] = df_expanded
            else:
                robot_print_info("Column 0 not found in DataFrame. Skipping filename split.")

            if 5 in df.columns:
                df_expanded_f = df[5].astype(str).str.split('.', expand=True)
                df[list(range(df.shape[1], df.shape[1] + df_expanded_f.shape[1]))] = df_expanded_f

                try:
                    df["H"] = df.iloc[:, -4] + "." + df.iloc[:, -2]
                    df["I"] = df["H"]
                except Exception as e:
                    robot_print_info(f"Error creating columns H and I: {e}")
            else:
                robot_print_info("Column 5 not found in DataFrame. Skipping dot split.")

            if "I" in df.columns:
                df_expanded_i = df["I"].astype(str).str.split('.', expand=True)
                df[list(range(df.shape[1], df.shape[1] + df_expanded_i.shape[1]))] = df_expanded_i
            else:
                robot_print_info("Column I not found. Skipping further split.")

            for idx, row in enumerate(df.values, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=idx, column=col_idx, value=value)

            updated_path = self.config_manager.get_blinking_parameters('update_file_path')
            wb.save(updated_path)
            robot_print_info(f"Formatted Excel saved at: {updated_path}")

        except Exception as e:
            robot_print_info(f"Exception in format_inside_excel: {e}")

    def create_new_sheet(self):
        """
        Create a new sheet in the Excel workbook specified in the config.

        The new sheet is populated with values from column A of the active sheet.
        """

        # file_path = self.config_manager.get_excel_path()
        file_path = self.config_manager.get_blinking_parameters('excel_path')
        wb = load_workbook(file_path)
        active_sheet = wb.active
        column_a_data = [cell.value for cell in active_sheet['A']]  # Extract column A data
        new_sheet = wb.create_sheet(title="NewSheet", index=wb.index(active_sheet) + 1)
        for row_index, value in enumerate(column_a_data, start=1):
            new_sheet.cell(row=row_index, column=1, value=value)
        wb.save(file_path)

    def extract_timestamp_from_filename(self, filename):
        """
        Extract a timestamp from the given filename based on known patterns.

        Supports two formats:
        - Format 1: image_YYYYMMDD_HH_MM_SS_micro.jpg
        - Format 2: -1-07d_00h_54m_14.046-img-1-

        Also resizes the image to 1920x720 if needed.

        Args:
            filename (str): Name of the image file.

        Returns:
            datetime or float or None: Parsed timestamp or total seconds, or None on failure.
        """

        try:
            # Format 1: image_YYYYMMDD_HH_MM_SS_micro.jpg
            if "image_" in filename:
                ts = filename.split("image_")[1].split(".jpg")[0]
                return datetime.strptime(ts, "%Y%m%d_%H_%M_%S_%f")

            # Format 2: -1-07d_00h_54m_14.046-img-1-
            match = re.search(r"(\d{2})d_(\d{2})h_(\d{2})m_(\d{2})\.(\d+)", filename)
            if match:
                days, hours, minutes, seconds, milliseconds = map(int, match.groups())
                total_seconds = days * 86400 + hours * 3600 + minutes * 60 + seconds + milliseconds / 1000.0

                primary_path = os.path.join(self.get_ivision_image_path(), filename)
                # fallback_path = os.path.join(self.config_manager.get_buffer_path(), filename)
                fallback_path = self.config_manager.get_blinking_parameters('buffer_path')

                image_path = primary_path if os.path.exists(primary_path) else fallback_path

                if os.path.exists(image_path):
                    img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
                    if img is not None:
                        if img.shape[0] != 720 or img.shape[1] != 1920:
                            resized_img = cv2.resize(img, (1920, 720))
                            cv2.imwrite(image_path, resized_img)
                            robot_print_info(f"Image '{filename}' resized to 1920x720")
                        else:
                            robot_print_info(f"Image '{filename}' already 1920x720, skipping resize")
                    else:
                        robot_print_info(f"Failed to read image '{filename}' for resizing")
                else:
                    robot_print_info(f"Image '{filename}' not found in both primary and buffer paths")

                return total_seconds

            raise ValueError("Unrecognized timestamp format")

        except Exception as e:
            robot_print_info(f"Timestamp parse error: {e}")
            return None

    def resize_image_if_needed(self, filename):
        """
        Resizes the image to 1920x720 if it's not already that size.

        Args:
            filename (str): Name of the image file.
        """
        try:
            primary_path = os.path.join(self.get_ivision_image_path(), filename)
            fallback_path = self.config_manager.get_blinking_parameters('buffer_path')
            image_path = primary_path if os.path.exists(primary_path) else os.path.join(fallback_path, filename)

            if os.path.exists(image_path):
                img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
                if img is not None:
                    if img.shape[0] != 720 or img.shape[1] != 1920:
                        resized_img = cv2.resize(img, (1920, 720))
                        cv2.imwrite(image_path, resized_img)
                        robot_print_info(f"Image '{filename}' resized to 1920x720")
                    else:
                        robot_print_info(f"Image '{filename}' already 1920x720, skipping resize")
                else:
                    robot_print_info(f"Failed to read image '{filename}' for resizing")
            else:
                robot_print_info(f"Image '{filename}' not found in both primary and buffer paths")
        except Exception as e:
            robot_print_info(f"Error resizing image '{filename}': {e}")

    def get_duration_in_seconds(self, start, end):
        """
        Calculate the duration in seconds between two timestamps.

        Args:
            start (datetime or float): Start time.
            end (datetime or float): End time.

        Returns:
            float: Duration in seconds.
        """

        if isinstance(start, datetime) and isinstance(end, datetime):
            return (end - start).total_seconds()
        return end - start

    def Calculate_All_Blink_Frequencies(self, valid_pairs, timestamps):
        """
        Calculate blink frequencies for all valid ON-OFF pairs.

        Args:
            valid_pairs (list): List of (on, off) index pairs.
            timestamps (list): List of datetime objects.

        Returns:
            list: List of blink frequencies in Hz.
        """

        try:
            frequencies = []
            for on, off in valid_pairs:
                start = min(on[0], off[0])
                end = max(on[1], off[1])
                duration = (timestamps[end] - timestamps[start]).total_seconds()
                if duration > 0:
                    frequencies.append(1.0 / duration)
            return frequencies
        except Exception as e:
            robot_print_info(f"Exception in Calculate_All_Blink_Frequencies: {e}")
            return []

    def Calculate_First_Blink_Frequency(self, valid_pairs, timestamps, first_off_end_index):
        """
        Calculate the frequency of the first blink after a given OFF index.

        Args:
            valid_pairs (list): List of (on, off) index pairs.
            timestamps (list): List of datetime objects.
            first_off_end_index (int): Index after which to find the first blink.

        Returns:
            float or None: Frequency in Hz, or None if not found.
        """

        try:
            for on, off in valid_pairs:
                if on[0] > first_off_end_index:
                    start = min(on[0], off[0])
                    end = max(on[1], off[1])
                    duration = timestamps[end] - timestamps[start]
                    if isinstance(duration, timedelta):
                        duration = duration.total_seconds()
                    if duration > 0:
                        return 0.5 / duration
                    break
            return None
        except Exception as e:
            robot_print_info(f"Exception in Calculate_First_Blink_Frequency: {e}")
            return None

    def preprocess_pattern_matches(self, test_case_name, files, threshold):
        """
        Run pattern matching for all files using the specified template and threshold.

        Args:
            test_case_name (str): Name of the test case.
            files (list): List of image filenames.
            threshold (float): Matching threshold.

        Returns:
            dict: Mapping of filenames to match results (True/False).
        """

        try:
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            template_path = self.config_manager.get_blinking_parameters('template_path')
            return {
                f: self.ivision_check_pattern_matching_for_blinking(
                    test_case_name,
                    os.path.join(buffer_path, f),
                    template_path,
                    threshold=float(threshold),
                    save_image=False,
                    save_path=None
                )
                for f in files
            }
        except Exception as e:
            robot_print_info(f"Exception in preprocess_pattern_matches: {e}")
            return {}

    def update_telltale_detection_status(self, test_case_name, sheet_name, match_results):
        """
        Updates detection status in an Excel sheet based on match results.

        Args:
            test_case_name (str): Name of the test case.
            sheet_name (str): Excel sheet to update.
            match_results (dict): Mapping of filenames to detection status (True/False).
        """

        try:
            excel_path = self.config_manager.get_blinking_parameters('update_file_path')

            wb = openpyxl.load_workbook(excel_path)
            ws = wb[sheet_name]

            for row in range(1, ws.max_row + 1):
                filename = ws.cell(row=row, column=1).value
                if not filename:
                    continue

                match = match_results.get(filename)
                if match is None:
                    ws.cell(row=row, column=2, value="File Not Found")
                else:
                    status = "TT Detected" if match else "TT Not Detected"
                    ws.cell(row=row, column=2, value=status)

            result_dir = self.get_ivision_result_path()
            os.makedirs(result_dir, exist_ok=True)
            result_excel_path = os.path.join(result_dir, f"telltale_detection_result_{test_case_name}.xlsx")
            wb.save(result_excel_path)
            wb.close()

            robot_print_info(f"Updated Excel saved at: {result_excel_path}")

        except Exception as e:
            robot_print_info(f"Exception in update_telltale_detection_status: {e}")

    def Blink_measurement(self, test_case_name, threshold):
        """
        Analyzes blinking pattern from image files for a given test case.

        This method processes image files from a specified buffer or input path,
        detects blinking sequences based on pattern matching, computes blink durations
        and frequencies, and saves the analysis results including visualizations and
        statistical data into an Excel sheet.

        Args:
            test_case_name (str): The name of the test case for analysis.
            threshold (float): The matching threshold for pattern detection.

        Returns:
            dict: A dictionary containing various statistical metrics related to blinking,
                  such as durations, rates, counts, and timestamps.
        """
        start_time = datetime.now()
        self.set_ivision_test_case_folder(test_case_name)

        buffer_path = self.config_manager.get_blinking_parameters('buffer_path')

        raw_value = self.config_manager.get_blinking_parameters("Input_Images_Required")

        input_images_required = raw_value.strip().lower() == "true"

        if input_images_required:
            robot_print_info("Copying files from buffer path...")
            copy_result = self.ivision_copy_all_files(buffer_path)
            robot_print_info(copy_result)
            image_path = self.get_ivision_image_path()
        else:
            image_path = buffer_path
            robot_print_info(f"Using buffer path directly: {buffer_path}")

        files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
        timestamps = [self.extract_timestamp_from_filename(f) for f in files]
        overall_time_period = self.get_duration_in_seconds(timestamps[0], timestamps[-1]) if timestamps else None

        match_results = self.preprocess_pattern_matches(test_case_name, files, threshold)

        first_on_index = None
        on_sequences = []
        off_sequences = []
        in_on = False
        in_off = False
        current_on_start = None
        current_off_start = None

        for i, f in enumerate(files):
            match = match_results[f]
            if match:
                if first_on_index is None:
                    first_on_index = i
                if not in_on:
                    current_on_start = i
                    in_on = True
                    if in_off:
                        off_sequences.append((current_off_start, i - 1))
                        in_off = False
            else:
                if not in_off:
                    current_off_start = i
                    in_off = True
                    if in_on:
                        on_sequences.append((current_on_start, i - 1))
                        in_on = False

        if in_on:
            on_sequences.append((current_on_start, len(files) - 1))
        if in_off:
            off_sequences.append((current_off_start, len(files) - 1))

        if first_on_index is not None:
            first_on_file = files[first_on_index]
            first_on_timestamp = timestamps[first_on_index]
            if isinstance(first_on_timestamp, datetime):
                first_on_timestamp_str = first_on_timestamp.strftime("%Y%m%d_%H%M%S")
            else:
                first_on_timestamp_str = f"{first_on_timestamp:.3f}".replace('.', '_')
            result_filename = f"first_on_result_{first_on_index}.png"
            result_folder = self.get_ivision_result_path()
            os.makedirs(result_folder, exist_ok=True)
            result_path = os.path.join(result_folder, result_filename)

            self.ivision_check_pattern_matching_for_blinking(
                test_case_name,
                os.path.join(image_path, first_on_file),
                self.config_manager.get_blinking_parameters('template_path'),
                threshold=float(threshold),
                save_image=True,
                save_path=result_path
            )

            robot_print_info(f"First ON result image saved at: {result_path}")

        first_off_end_index = off_sequences[0][1] if off_sequences else None
        last_on_start_index = on_sequences[-1][0] if on_sequences else None

        valid_pairs = []
        for on in on_sequences:
            if (first_off_end_index is not None and on[0] > first_off_end_index) and \
                    (last_on_start_index is not None and on[0] < last_on_start_index):
                matching_off = next((off for off in off_sequences if off[0] > on[1] and off[0] < last_on_start_index),
                                    None)
                if matching_off:
                    valid_pairs.append((on, matching_off))

        blink_durations = []
        for on, off in valid_pairs:
            start = min(on[0], off[0])
            end = max(on[1], off[1])
            # duration = self.get_duration_in_seconds(timestamps[start], timestamps[end])
            # duration = (timestamps[end] - timestamps[start]).total_seconds()
            duration = timestamps[end] - timestamps[start]
            if isinstance(duration, timedelta):
                duration = duration.total_seconds()

            blink_durations.append(duration)

        total_blinks = len(blink_durations)
        average_blink_period = sum(blink_durations) / total_blinks if total_blinks else None
        max_blink_duration = max(blink_durations) if blink_durations else None
        min_blink_duration = min(blink_durations) if blink_durations else None

        blink_frequencies = []
        try:
            for duration in blink_durations:
                if duration > 0:
                    blink_frequencies.append(1.0 / duration)
            average_blink_frequency = sum(blink_frequencies) / len(blink_frequencies) if blink_frequencies else None
        except Exception as e:
            robot_print_info(f"Error calculating iVision_Blink frequencies: {e}")
            blink_frequencies = []
            average_blink_frequency = None

        first_blink_frequency = 1.0 / blink_durations[0] if blink_durations and blink_durations[0] > 0 else None
        first_blink_rate = self.Calculate_First_Blink_Frequency(valid_pairs, timestamps,
                                                                first_off_end_index)

        filtered_on = [seq for seq in on_sequences if seq[0] > first_off_end_index and seq[0] < last_on_start_index]
        filtered_off = [seq for seq in off_sequences if seq[0] > first_off_end_index and seq[0] < last_on_start_index]

        pre_detection_duration = self.get_duration_in_seconds(
            timestamps[0], timestamps[first_on_index]
        ) if first_on_index is not None else None

        post_detection_duration = self.get_duration_in_seconds(timestamps[last_on_start_index], timestamps[
            -1]) if last_on_start_index is not None else None

        start_to_first_on_after_first_off = self.get_duration_in_seconds(
            timestamps[0], timestamps[filtered_on[0][0]]
        ) if filtered_on else None


        durations = {
            "start_to_first_on": self.get_duration_in_seconds(timestamps[0],
                                                              timestamps[first_on_index]) if first_on_index else None,
            # "start_to_first_on_after_first_off": start_to_first_on_after_first_off,
            "first_on_duration": self.get_duration_in_seconds(timestamps[filtered_on[0][0]],
                                                              timestamps[filtered_on[0][1]]) if filtered_on else None,
            "first_off_duration": self.get_duration_in_seconds(timestamps[filtered_off[0][0]], timestamps[
                filtered_off[0][1]]) if filtered_off else None,
            "first_full_blink_rate": first_blink_frequency if first_blink_frequency else None,
            "first_half_blink_rate": first_blink_rate if first_blink_rate else None,
            "Average_blink_time": average_blink_period if average_blink_period else None,
            "Blink_rate": average_blink_frequency if average_blink_frequency else None,
            "total_on_sequence_occurred": len(on_sequences),
            "on_sequence_count": len(filtered_on),
            "total_off_sequence_occurred": len(off_sequences),
            "off_sequence_count": len(filtered_off),
            "max_blink_duration": max_blink_duration,
            "min_blink_duration": min_blink_duration,
            # "Blink_Rate": 1 / ((
            #                                max_blink_duration + min_blink_duration) / 2) if max_blink_duration and min_blink_duration else None,
            "total_blinks_detected": total_blinks,
            "first_on_time_occurred_at": ((timestamps[filtered_on[0][0]] - timestamps[0]).total_seconds() if isinstance(
                timestamps[filtered_on[0][0]], datetime) else
                                          timestamps[filtered_on[0][0]] - timestamps[0]) if filtered_on else None,

            "first_off_time_occurred_at": (
                (timestamps[filtered_off[0][0]] - timestamps[0]).total_seconds()
                if isinstance(timestamps[filtered_off[0][0]], datetime) else
                timestamps[filtered_off[0][0]] - timestamps[0]
            ) if filtered_off else None,

            "overall_time_period_seconds": overall_time_period,
            "pre_detection_duration": pre_detection_duration,
            "post_detection_duration": post_detection_duration,
            "blink_summary": f"In {int(overall_time_period)} seconds, {total_blinks} blinks were detected." if overall_time_period else None
        }

        self.save_files_to_excel()
        self.create_new_sheet()
        self.format_inside_excel()
        self.update_telltale_detection_status(test_case_name, "NewSheet", match_results)

        return "\n".join(f"{k}: {v}" for k, v in durations.items())

    def check_template_match(self, test_case_name, threshold):
        """
        Checks if any image in the buffer matches the template.

        Returns True if a match is found, otherwise False.
        """
        try:
            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            # input_images_required = self.config_manager.get_blinking_parameters(
            #     "Input_Images_Required").strip().lower() == "true"
            #
            # if input_images_required:
            #     robot_print_info("Copying files from buffer path...")
            #     copy_result = self.ivision_copy_all_files(buffer_path)
            #     robot_print_info(copy_result)
            #     image_path = self.get_ivision_image_path()
            # else:
            image_path = buffer_path
            robot_print_info(f"Using buffer path directly: {buffer_path}")

            if not os.path.exists(image_path):
                robot_print_info(f"Image path does not exist: {image_path}")
                return False

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            template_path = self.config_manager.get_blinking_parameters('template_path')
            result_folder = self.get_ivision_result_path()
            os.makedirs(result_folder, exist_ok=True)

            for f in files:
                image_full_path = os.path.join(image_path, f)
                match = self.ivision_check_pattern_matching_for_blinking(
                    test_case_name=test_case_name,
                    input_image_path=image_full_path,
                    template_image_path=template_path,
                    threshold=float(threshold),
                    save_image=True,
                    save_path=os.path.join(result_folder, f"matched_{f}")
                )
                if match:
                    robot_print_info(f"Template matched in image: {f}")
                    return True

            robot_print_info("No template match found in any image.")
            return False

        except Exception as e:
            robot_print_info(f"Error in check_template_found: {e}")
            return False

    def measure_start_to_first_template_match_time(self, test_case_name, threshold):
        """
        Measures time from function start to the first template match.

        Returns the duration in seconds, or None if no match is found.
        """
        try:
            start_time = datetime.now()
            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')

            image_path = buffer_path
            robot_print_info(f"Using buffer path directly: {buffer_path}")

            if not os.path.exists(image_path):
                robot_print_info(f"Image path does not exist: {image_path}")
                return None

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            template_path = self.config_manager.get_blinking_parameters('template_path')

            for f in files:
                image_full_path = os.path.join(image_path, f)
                self.resize_image_if_needed(f)

                match = self.ivision_check_pattern_matching_for_blinking(
                    test_case_name=test_case_name,
                    input_image_path=image_full_path,
                    template_image_path=template_path,
                    threshold=float(threshold),
                    save_image=False,
                    save_path=None
                )
                if match:
                    end_time = datetime.now()
                    duration = (end_time - start_time).total_seconds()
                    robot_print_info(f"Template matched in image: {f} after {duration:.2f} seconds")
                    return duration

            robot_print_info("No template match found in any image.")
            return None
        except Exception as e:
            robot_print_info(f"Error in measure_start_to_template: {e}")
            return None

    def measure_first_consecutive_template_match_time(self, test_case_name, threshold):
        """
        Measures the duration of the first consecutive ON (template match) sequence.

        Returns both the ON sequence duration and the total time from function start to the end of that sequence.
        """
        try:

            start_time = datetime.now()  # Start of function execution

            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            # input_images_required = self.config_manager.get_blinking_parameters(
            #     "Input_Images_Required").strip().lower() == "true"
            #
            # if input_images_required:
            #     robot_print_info("Copying files from buffer path...")
            #     copy_result = self.ivision_copy_all_files(buffer_path)
            #     robot_print_info(copy_result)
            #     image_path = self.get_ivision_image_path()
            # else:
            image_path = buffer_path
            # robot_print_info(f"Using buffer path directly: {buffer_path}")

            if not os.path.exists(image_path):
                robot_print_info(f"Image path does not exist: {image_path}")
                return None

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            timestamps = [self.extract_timestamp_from_filename(f) for f in files]
            template_path = self.config_manager.get_blinking_parameters('template_path')

            in_on = False
            on_start_index = None
            on_end_index = None

            for i, f in enumerate(files):
                image_full_path = os.path.join(image_path, f)
                match = self.ivision_check_pattern_matching_for_blinking(
                    test_case_name=test_case_name,
                    input_image_path=image_full_path,
                    template_image_path=template_path,
                    threshold=float(threshold),
                    save_image=False,
                    save_path=None
                )

                if match:
                    if not in_on:
                        on_start_index = i
                        in_on = True
                    on_end_index = i
                elif in_on:
                    break

            if on_start_index is not None and on_end_index is not None:
                on_duration = self.get_duration_in_seconds(timestamps[on_start_index], timestamps[on_end_index])
                total_duration = (datetime.now() - start_time).total_seconds()
                robot_print_info(f"First ON sequence duration: {on_duration:.2f} seconds")
                robot_print_info(f"Total time from function start to end of ON sequence: {total_duration:.2f} seconds")
                return on_duration

            robot_print_info("No consecutive ON sequence found.")
            return None

        except Exception as e:
            robot_print_info(f"Error in measure_first_consecutive_template_found: {e}")
            return None

    def measure_last_template_match_to_end_time(self, test_case_name, threshold):
        """
        Calculates the duration from the last ON sequence before the last OFF sequence to the end of that OFF.

        Returns the duration in seconds, or None if no valid ON/OFF sequence is found.
        """
        try:
            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            image_path = buffer_path
            robot_print_info(f"Using buffer path directly: {buffer_path}")

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            timestamps = [self.extract_timestamp_from_filename(f) for f in files]
            match_results = self.preprocess_pattern_matches(test_case_name, files, threshold)

            on_sequences, off_sequences = self.get_on_off_sequences(files, match_results)

            if not off_sequences:
                robot_print_info("No OFF sequence found.")
                return None
            last_off_start, last_off_end = off_sequences[-1]

            last_on_before_off = None
            for on_seq in reversed(on_sequences):
                if on_seq[1] < last_off_start:
                    last_on_before_off = on_seq
                    break

            if last_on_before_off:
                start_index = last_on_before_off[0]
                end_index = last_off_end
                duration = self.get_duration_in_seconds(timestamps[start_index], timestamps[end_index])
                robot_print_info(f"Duration from last ON before last OFF to end of last OFF: {duration:.3f} seconds")
                return duration
            else:
                robot_print_info("No ON sequence found before the last OFF.")
                return None

        except Exception as e:
            robot_print_info(f"Error in measure_last_on_before_last_off_duration: {e}")
            return None

    def number_of_continous_template_match_found(self, test_case_name, threshold):
        """
        Counts the number of continuous ON (template match) sequences in the image set.

        Returns the total number of ON sequences detected.
        """
        try:
            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            input_images_required = self.config_manager.get_blinking_parameters(
                "Input_Images_Required").strip().lower() == "true"

            if input_images_required:
                robot_print_info("Copying files from buffer path...")
                copy_result = self.ivision_copy_all_files(buffer_path)
                robot_print_info(copy_result)
                image_path = self.get_ivision_image_path()
            else:
                image_path = buffer_path
                robot_print_info(f"Using buffer path directly: {buffer_path}")

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]

            for f in files:
                self.resize_image_if_needed(f)

            match_results = self.preprocess_pattern_matches(test_case_name, files, threshold)

            on_sequences, _ = self.get_on_off_sequences(files, match_results)
            self.save_files_to_excel()
            self.create_new_sheet()
            self.format_inside_excel()
            self.update_telltale_detection_status(test_case_name, "NewSheet", match_results)

            return len(on_sequences)
        except Exception as e:
            robot_print_info(f"Error in average_template_found: {e}")
            return 0

    def average_template_match_sequence_duration(self, test_case_name, threshold):
        """
        Calculates the average duration of all ON (template match) sequences.

        Returns the average ON duration in seconds, or 0 if no ON sequences are found.
        """
        try:
            self.set_ivision_test_case_folder(test_case_name)
            buffer_path = self.config_manager.get_blinking_parameters('buffer_path')
            input_images_required = self.config_manager.get_blinking_parameters(
                "Input_Images_Required").strip().lower() == "true"

            if input_images_required:
                robot_print_info("Copying files from buffer path...")
                copy_result = self.ivision_copy_all_files(buffer_path)
                robot_print_info(copy_result)
                image_path = self.get_ivision_image_path()
            else:
                image_path = buffer_path
                robot_print_info(f"Using buffer path directly: {buffer_path}")

            files = [f for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f))]
            timestamps = [self.extract_timestamp_from_filename(f) for f in files]
            match_results = self.preprocess_pattern_matches(test_case_name, files, threshold)

            on_sequences, _ = self.get_on_off_sequences(files, match_results)

            durations = []
            for start_idx, end_idx in on_sequences:
                duration = self.get_duration_in_seconds(timestamps[start_idx], timestamps[end_idx])
                durations.append(duration)

            if durations:
                average_duration = sum(durations) / len(durations)
                robot_print_info(f'{durations}')
                robot_print_info(f"Average ON sequence duration: {average_duration:.3f} seconds")
                self.save_files_to_excel()
                self.create_new_sheet()
                self.format_inside_excel()
                self.update_telltale_detection_status(test_case_name, "NewSheet", match_results)

                return average_duration

            else:
                robot_print_info("No ON sequences found.")
                return 0

        except Exception as e:
            robot_print_info(f"Error in average_on_sequence_duration: {e}")
            return 0

    def get_on_off_sequences(self, files, match_results):
        """
        Identifies ON and OFF sequences based on template match results.

        Returns two lists of tuples: ON sequences and OFF sequences with their start and end indices.
        """
        try:
            on_sequences = []
            off_sequences = []
            in_on = False
            in_off = False
            current_on_start = None
            current_off_start = None

            for i, f in enumerate(files):
                match = match_results[f]
                if match:
                    if not in_on:
                        current_on_start = i
                        in_on = True
                        if in_off:
                            off_sequences.append((current_off_start, i - 1))
                            in_off = False
                else:
                    if not in_off:
                        current_off_start = i
                        in_off = True
                        if in_on:
                            on_sequences.append((current_on_start, i - 1))
                            in_on = False

            if in_on:
                on_sequences.append((current_on_start, len(files) - 1))
            if in_off:
                off_sequences.append((current_off_start, len(files) - 1))

            return on_sequences, off_sequences
        except Exception as e:
            robot_print_info(f"Error in get_on_off_sequences: {e}")
            return [], []
