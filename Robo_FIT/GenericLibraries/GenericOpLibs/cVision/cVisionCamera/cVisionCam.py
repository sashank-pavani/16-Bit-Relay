import cv2
import os
import time
import sys
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import (
    robot_print_error, robot_print_debug, robot_print_warning, robot_print_info
)
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass
from Robo_FIT.GenericLibraries.GenericOpLibs.cVision.cVisionCamera.ConfigurationManager import ConfigurationManager
import time
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import pandas as pd
from Robo_FIT.GenericLibraries.GenericOpLibs.UITesting.ImageFinder import ImageFinder
import threading
import ctypes


class cVisionCam:

    def __init__(self):
        """Initializes the cVision class."""
        try:
            self.common_keyword = CommonKeywordsClass()
            self.test_case_name = None
            self.config_manager = ConfigurationManager()
            self.image = ImageFinder()
            robot_print_info("cVision initialized.")
        except Exception as e:
            robot_print_error(f"Initialization failed: {str(e)}")
            raise

    def cVision_WebCam_Capture(self, filename: str):
        """Captures an image from the camera and saves it.

                Args:
                    filename (str): The name of the file to save the image.

                Returns:
                    str: The path of the saved image file.
        """
        try:
            cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            ret, frame = cap.read()

            if not ret:
                robot_print_error("Error with the camera. Image capture failed.")
                cap.release()
                return None

            Pname = os.path.basename(sys.argv[0])[0:4]
            robot_print_info(f"Pname: {Pname}")

            filename = filename + ".jpg"
            file_name = os.path.join(self.get_cvision_image_path(), filename)
            robot_print_info(f"file_name: {file_name}")

            cv2.imwrite(file_name, frame)
            robot_print_info("Screenshot done!")

            cap.release()
            cv2.destroyAllWindows()
            robot_print_info(f"Camera Image saved at: {file_name}")
            return file_name
        except Exception as e:
            robot_print_error(f"Failed to capture image: {str(e)}")
            return None

    def set_cVision_test_case_folder(self, test_case_name: str):
        """Sets the test case name for the instance.

                        Args:
                            test_case_name (str): The name of the test case.
        """

        try:
            self.test_case_name = test_case_name
            robot_print_info(f"Test case name set to: {self.test_case_name}")
        except Exception as e:
            robot_print_error(f"Error setting test case name: {str(e)}")

    def get_cvision_image_path(self):
        """Retrieves the path to the cVision image directory.

                Returns:
                    str: The path to the cVision image directory.

                Raises:
                    ValueError: If test_case_name is not set.
                    OSError: If directory creation fails.
        """
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")

            subfolders = [self.config_manager.get_cvision_base_folder(),
                          self.config_manager.get_cvision_image_subfolder(), self.test_case_name]
            image_path = os.path.join(self.common_keyword.get_report_path(), *subfolders)

            if not os.path.isdir(image_path):
                os.makedirs(image_path, mode=0o777)
                time.sleep(1)
            return image_path
        except OSError as os_err:
            robot_print_error(f"Error creating the IMAGE directory: {os_err}")
            raise
        except Exception as e:
            robot_print_error(f"Unexpected error in get_cvision_image_path: {str(e)}")
            raise

    def get_cvision_result_path(self):
        """Retrieves the path to the cVision result directory.

                       Returns:
                           str: The path to the cVision result directory.

                       Raises:
                           ValueError: If test_case_name is not set.
                           OSError: If directory creation fails.
        """
        try:
            if self.test_case_name is None:
                raise ValueError("Test case name is not set.")

            subfolders = [self.config_manager.get_cvision_base_folder(),
                          self.config_manager.get_cvision_result_subfolder(), self.test_case_name]
            result_path = os.path.join(self.common_keyword.get_report_path(), *subfolders)

            if not os.path.isdir(result_path):
                os.makedirs(result_path, mode=0o777)
                time.sleep(1)
            return result_path
        except OSError as os_err:
            robot_print_error(f"Error creating the RESULT directory: {os_err}")
            raise
        except Exception as e:
            robot_print_error(f"Unexpected error in get_cvision_result_path: {str(e)}")
            raise

    def get_camera_fps(self, camera_index=0, num_frames=1500):
        """ This function is used to get the FPS of given camera
        Argument:
        camera_index - Index for all connected cameras. 0 for default camera
        """
        cap = cv2.VideoCapture(camera_index)
        if not cap.isOpened():
            print("Error: Could not open camera.")
            return
        start_time = time.time()
        for _ in range(num_frames):
            ret, frame = cap.read()
            if not ret:
                print("Error: Could not read frame.")
                break
        end_time = time.time()
        elapsed_time = end_time - start_time
        fps = num_frames / elapsed_time
        cap.release()
        return round(fps)

    def cVision_fps_Cam_Capture(self,  cvision_path, time_limit=5, fps_required=10, camera=0, save_dir="captured_images"):
        """ This function is used to capture continuous images for the given time at given FPS.

        Arguments:
            cvision_path - this is the return value of function get_cvision_result_path
            time_limit - How much time in seconds you want to capture
            fps_required - user defined FPS
            camera - Camera index ( optional )
            save_dir - Directory name insise reports folder ( optional )
            """
        robot_print_info(f"Entered cVision_fps_Cam_Capture")
        if self.config_manager.get_manual_fps() == "Yes" or self.config_manager.get_manual_fps() == "yes":
            fps = self.config_manager.get_fps()
            robot_print_info(f"Manual FPS taken from JSON")
        else:
            fps = self.get_camera_fps(camera)
            robot_print_info(f"FPS is calculated dynamically")
        buffer_limit = time_limit * fps
        robot_print_info(f"buffer_limit == {buffer_limit}")
        seconds_per_frame = fps // fps_required
        robot_print_info(f"seconds_per_frame === {seconds_per_frame}")
        #save_dir = os.path.join(self.get_cvision_image_path(), save_dir)
        save_dir_image = os.path.join(cvision_path, save_dir)
        robot_print_info(f"save_dir_image === {save_dir_image}")
        if not os.path.exists(save_dir_image):
            os.makedirs(save_dir_image)
        cap = cv2.VideoCapture(camera)
        if not cap.isOpened():
            print("Error: Could not open webcam.")
            return
        print(f"Capturing images... ")
        image_count = 0

        while image_count < buffer_limit:
            ret, frame = cap.read()
            if not ret:
                print("Error: Could not read frame.")
                break
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            file_name = f"{save_dir_image}/image_{timestamp}.jpg"

            # Save the frame as an image
            if not image_count % seconds_per_frame:
                cv2.imwrite(file_name, frame)
            image_count += 1

        cap.release()
        cv2.destroyAllWindows()
        print(f"Done! {image_count} images saved in '{save_dir_image}'.")
        files = self.list_files_in_folder(save_dir_image)
        excel_path = os.path.join(cvision_path, "cVision_img.xlsx")
        robot_print_info(f"Excel path == {excel_path}")
        self.save_files_to_excel(excel_path, files)
        self.format_inside_excel(excel_path)

        #time.sleep(30)
        #print("Wait time completed")

    def list_files_in_folder(self, get_destination_path):
        """This function is created to list all files in a folder.
                return : list of files inside the folder
                """
        try:
            files = os.listdir(get_destination_path)
            files = [f for f in files if os.path.isfile(os.path.join(get_destination_path, f))]
            return files
        except Exception as e:
            return str(e)

    def save_files_to_excel(self, get_excel_path, files):

        """This function is created to copy all folders from one folder to other .
            parm: files - List of image names to be processed
            excel_path - path in CRE - External files folder to save the excel sheet.

            Note: please provide full path that includes excel workbook name.
            System will erase existing data if excel sheet is available or
            it will create new excel workbook in case it is not available.
            Close or Delete the workbook befor running the script to avoide errors.

            return : success message
                                """
        try:

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for index, file_name in enumerate(files, start=1):
                sheet[f"A{index}"] = file_name
            workbook.save(get_excel_path)
            return f"Excel file created and saved at {get_excel_path}"
        except Exception as e:
            return str(e)

    def format_inside_excel(self, file_path):
        """ This function is used to perform data formating inside the excel workbook mentioned in JSON file.
        No argument required"""

        self.create_new_sheet(file_path)
        robot_print_info(f"get_excel_path{file_path}")
        wb = load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=1):
            for cell in row:
                ws.cell(row=cell.row, column=2, value=cell.value)
        df = pd.DataFrame(ws.values)
        df_expanded = df[0].str.split('_', expand=True)
        df[list(range(df_expanded.shape[1]))] = df_expanded

        df_expanded_f = df[3].str.split('.', expand=True)
        df[list(range(df.shape[1], df.shape[1] + df_expanded_f.shape[1]))] = df_expanded_f
        df[6] = df[2] + "." + df[4]
        for idx, row in enumerate(df.values, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=idx, column=col_idx, value=value)
        wb.save(file_path)

    def create_new_sheet(self, file_path):
        """ This function is used to create a new sheet inside the excel workbook mentioned in JSON file
        No argument required"""
        wb = load_workbook(file_path)
        active_sheet = wb.active
        column_a_data = [cell.value for cell in active_sheet['A']]  # Extract column A data
        new_sheet = wb.create_sheet(title="NewSheet", index=wb.index(active_sheet) + 1)
        for row_index, value in enumerate(column_a_data, start=1):
            new_sheet.cell(row=row_index, column=1, value=value)
        wb.save(file_path)

    def cVision_Async_fps_Cam_Capture(self,cvision_path, time_limit=5, fps_required=10, camera=0, save_dir="captured_images"):
        """ This function is defined to simultaniouly run cVision_fps_Cam_Capture  while executing other activities"""
        self.thread1 = threading.Thread(target=self.cVision_fps_Cam_Capture, args=(cvision_path, time_limit, fps_required, camera,  save_dir))
        self.thread1.start()
    def cVision_Async_stop_fps_Cam_Capture(self):
        """ This function is used to stop the thread created by cVision_Async_fps_Cam_Capture """
        thread = self.thread1
        if not thread.is_alive():
            return

        thread_id = ctypes.c_long(thread.ident)
        res = ctypes.pythonapi.PyThreadState_SetAsyncExc(thread_id, ctypes.py_object(SystemExit))
        if res == 0:
            raise ValueError("Invalid thread ID")
        elif res > 1:
            ctypes.pythonapi.PyThreadState_SetAsyncExc(thread_id, None)
            raise SystemError("PyThreadState_SetAsyncExc failed")

if __name__ =='__main__':
    excel_test_path = "C:\RF_Env\Dhivya\RoboFIT_DI_RE_WC_RawCan_\CRE\SWE5_SWIntegrationTest\Test_Reports\RE_P3F2_OAT_Feb_19_2025_14_46_30\cVision\cVision_Result\TC4\cVision_img.xlsx"
    h = cVisionCam()
    h.format_inside_excel(excel_test_path)
