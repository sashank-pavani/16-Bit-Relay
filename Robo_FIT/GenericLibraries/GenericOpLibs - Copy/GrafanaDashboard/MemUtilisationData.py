import openpyxl

from Robo_FIT.GenericLibraries.GenericOpLibs.GrafanaDashboard.DataPost import DataPost
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.Constants import *
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.ProjectConfigManager import ProjectConfigManager
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CustomPrint import robot_print_error, robot_print_info
import os
from Robo_FIT.GenericLibraries.GenericOpLibs.TestArtifacts.CommonKeywordsClass import CommonKeywordsClass


class MemUtilisationData:
    def __init__(self):

        self.project_config_file = ProjectConfigManager()
        self.base_url = self.project_config_file.get_grafana_base_url()
        self.helper = CommonKeywordsClass()
        self.upload_to_grafana = DataPost()

    @property
    def android_free_excel_file_path(self):
        """
        This function gets excel file path of Android free
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_ANDROID_FREE_EXCEL_NAME
        )
        return file_path

    @property
    def android_mem_excel_file_path(self):
        """
        This function gets excel file path of Android mem
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_ANDROID_MEM_EXCEL_NAME
        )
        return file_path

    @property
    def android_procrank_excel_file_path(self):
        """
        This function gets excel file path of Android procrank
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_ANDROID_PROCRANK_EXCEL_NAME
        )
        return file_path

    @property
    def android_top_excel_file_path(self):
        """
        This function gets excel file path of Android top
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_ANDROID_TOP_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_pmem_excel_file_path(self):
        """
        This function gets excel file path of qnx pmem
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_QNX_PMEM_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_showmem_excel_file_path(self):
        """
        This function gets excel file path of qnx showmem
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_QNX_SHOWMEM_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_dma_excel_file_path(self):
        """
        This function gets excel file path of qnx dma
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_QNX_DMA_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_summary_excel_file_path(self):
        """
        This function gets excel file path of qnx summary
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_QNX_SUMMARY_EXCEL_NAME
        )
        return file_path

    @property
    def qnx_top_excel_file_path(self):
        """
        This function gets excel file path of qnx top
        :return: path of excel file
        """
        file_path = os.path.join(
            self.helper.get_report_path(),
            PERFORMANCE_UTILISATION_DIR,
            ROBO_MEMORY_DIR_NAME,
            ROBO_QNX_TOP_EXCEL_NAME
        )
        return file_path

    @property
    def upload_url(self):
        """
        This function returns the API endpoint URL for BOOT KPI
        :return: URL
        """
        return self.base_url + MEMORY_URL

    def _check_memory_file(self, file_type: str):
        """
        This function checks whether excel file path exists, if it does, it checks that sheet is not blank
        :return: True if excel file exists and is not blank else false
        """
        if file_type == ROBO_ANDROID_FREE_FILE_NAME:
            path = self.android_free_excel_file_path
            sheet = self._get_sheet_name(ROBO_ANDROID_FREE_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Android free data", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Boot KPI")
                return False
        elif file_type == ROBO_ANDROID_MEM_FILE_NAME:
            path = self.android_mem_excel_file_path
            sheet = self._get_sheet_name(ROBO_ANDROID_MEM_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Android mem data", print_in_report=True,
                                     underline=True)
                    return True
                else:
                    robot_print_info("No file generated for memory")
                    return False
        elif file_type == ROBO_ANDROID_TOP_FILE_NAME:
            path = self.android_top_excel_file_path
            sheet = self._get_sheet_name(ROBO_ANDROID_TOP_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Android top data", print_in_report=True,
                                     underline=True)
                    return True
                else:
                    robot_print_info("No file generated for memory")
                    return False

        elif file_type == ROBO_ANDROID_PROCRANK_FILE_NAME:
            path = self.android_procrank_excel_file_path
            sheet = self._get_sheet_name(ROBO_ANDROID_PROCRANK_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for Android procrank data", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Memory")
                return False
            
        elif file_type == ROBO_QNX_PMEM_FILE_NAME:
            path = self.qnx_pmem_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_PMEM_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for qnx pmem data", print_in_report=True,
                                     underline=True)
                    return True
                else:
                    robot_print_info("No file generated for memory")
                    return False
        elif file_type == ROBO_QNX_DMA_FILE_NAME:
            path = self.qnx_dma_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_DMA_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for qnx_dma data", print_in_report=True,
                                     underline=True)
                    return True
                else:
                    robot_print_info("No file generated for memory")
                    return False

        elif file_type == ROBO_QNX_SHOWMEM_FILE_NAME:
            path = self.qnx_showmem_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_SHOWMEM_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for qnx showmem data", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Memory")
                return False
            
        elif file_type == ROBO_QNX_SUMMARY_FILE_NAME:
            path = self.qnx_summary_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_SUMMARY_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for qnx summary data", print_in_report=True,
                                     underline=True)
                    return True
                else:
                    robot_print_info("No file generated for memory")
                    return False

        elif file_type == ROBO_QNX_TOP_FILE_NAME:
            path = self.qnx_top_excel_file_path
            sheet = self._get_sheet_name(ROBO_QNX_TOP_FILE_NAME)
            if os.path.exists(path):
                if self._is_sheet_not_blank(path, sheet):
                    robot_print_info(f"File generated and sheet is present for qnx top data", print_in_report=True,
                                     underline=True)
                    return True
            else:
                robot_print_info("No file generated for Memory")
                return False

    def _is_sheet_not_blank(self, excel_file_path, sheet_name):
        """
        This function checks that excel file is not blank
        :return:True if file is not blank else false
        """
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        workbook.close()
                        print("File is not empty")
                        return True
            workbook.close()
            robot_print_info("File is empty")
            return False
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")
            return False

    def _get_sheet_name(self, file_type: str):
        """
        This function gets name of sheet in excel file
        :return: Sheet name
        """
        try:
            if file_type == ROBO_ANDROID_FREE_FILE_NAME:
                excel_file_path = self.android_free_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_ANDROID_MEM_FILE_NAME:
                excel_file_path = self.android_mem_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_ANDROID_TOP_FILE_NAME:
                excel_file_path = self.android_top_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_ANDROID_PROCRANK_FILE_NAME:
                excel_file_path = self.android_procrank_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_QNX_PMEM_FILE_NAME:
                excel_file_path = self.qnx_pmem_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_QNX_SHOWMEM_FILE_NAME:
                excel_file_path = self.qnx_showmem_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_QNX_DMA_FILE_NAME:
                excel_file_path = self.qnx_dma_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_QNX_SUMMARY_FILE_NAME:
                excel_file_path = self.qnx_summary_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            elif file_type == ROBO_QNX_TOP_FILE_NAME:
                excel_file_path = self.qnx_top_excel_file_path
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames
                workbook.close()
                robot_print_info(sheet_names[0])
                return sheet_names[0]
            
        except Exception as e:
            robot_print_error(f"An error occurred: {e}")

    def __upload_android_free_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_ANDROID_FREE_FILE_NAME):
                android_mem_data_excel_file = self.android_free_excel_file_path
                android_mem_data_sheet_name = self._get_sheet_name(android_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(android_mem_data_excel_file,
                                                                         android_mem_data_sheet_name)

                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_android_mem_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_ANDROID_MEM_FILE_NAME):
                android_mem_data_excel_file = self.android_mem_excel_file_path
                android_mem_data_sheet_name = self._get_sheet_name(android_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(android_mem_data_excel_file,
                                                                         android_mem_data_sheet_name)

                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_android_top_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_ANDROID_TOP_FILE_NAME):
                android_mem_data_excel_file = self.android_top_excel_file_path
                android_mem_data_sheet_name = self._get_sheet_name(android_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(android_mem_data_excel_file,
                                                                         android_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_android_procrank_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """

        try:
            if self._check_memory_file(ROBO_ANDROID_PROCRANK_FILE_NAME):
                android_mem_data_excel_file = self.android_procrank_excel_file_path
                android_mem_data_sheet_name = self._get_sheet_name(android_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(android_mem_data_excel_file,
                                                                         android_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_qnx_dma_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_QNX_DMA_FILE_NAME):
                qnx_mem_data_excel_file = self.qnx_dma_excel_file_path
                qnx_mem_data_sheet_name = self._get_sheet_name(ROBO_QNX_DMA_FILE_NAME)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_mem_data_excel_file,
                                                                         qnx_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_qnx_pmem_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_QNX_PMEM_FILE_NAME):
                qnx_mem_data_excel_file = self.qnx_pmem_excel_file_path
                qnx_mem_data_sheet_name = self._get_sheet_name(qnx_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_mem_data_excel_file,
                                                                         qnx_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_qnx_showmem_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_QNX_SHOWMEM_FILE_NAME):
                qnx_mem_data_excel_file = self.qnx_dma_excel_file_path
                qnx_mem_data_sheet_name = self._get_sheet_name(qnx_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_mem_data_excel_file,
                                                                         qnx_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_qnx_summary_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_QNX_SUMMARY_FILE_NAME):
                qnx_mem_data_excel_file = self.qnx_summary_excel_file_path
                qnx_mem_data_sheet_name = self._get_sheet_name(qnx_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_mem_data_excel_file,
                                                                         qnx_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")

    def __upload_qnx_top_data(self):
        """
        This function is currently private as API is not implemented on Grafana server side
        It is implemented for future use and requires testing
        """
        try:
            if self._check_memory_file(ROBO_QNX_TOP_FILE_NAME):
                qnx_mem_data_excel_file = self.qnx_top_excel_file_path
                qnx_mem_data_sheet_name = self._get_sheet_name(qnx_mem_data_excel_file)
                data_json = self.upload_to_grafana.convert_excel_to_json(qnx_mem_data_excel_file,
                                                                         qnx_mem_data_sheet_name)
                data_payload = {
                    "url": self.upload_url,
                    "data_json": data_json
                }
                self.upload_to_grafana.post_data(**data_payload)
            else:
                robot_print_info("File not present. Please check report")
        except Exception as e:
            robot_print_error(f"An error occurred during Memory data upload: {e}")
